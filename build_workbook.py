#!/usr/bin/env python3
"""
Build a single .xlsx workbook with one tab per domain from the CSVs produced by
keyword_research.py. Uploading this to Google Drive with conversion enabled
yields one Google Sheet with two tabs.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
OUT_PATH = Path(__file__).parent / "Keyword Research.xlsx"

# (csv filename, tab name)
TABS = [
    ("sundaycitizen_co.csv",     "sundaycitizen.co"),
    ("brooklinen_com.csv",       "brooklinen.com"),
    ("potterybarn_com.csv",      "potterybarn.com"),
    ("latestbedding_com.csv",    "latestbedding.com"),
    ("tempurpedic_com.csv",      "tempurpedic.com"),
    ("us_pigletinbed_com.csv",   "pigletinbed.com"),
    ("parachutehome_com.csv",    "parachutehome.com"),
    ("brooklinen_expansion.csv", "brooklinen — expansion"),
    ("serp_competitors.csv",     "competing domains"),
    ("winnable_pages.csv",       "WINNABLE (100-1500)"),
    ("page_plan.csv",            "PAGE PLAN"),
    ("keyword_gap_pages.csv",    "GAP — pages (detail)"),
    ("keyword_gap.csv",          "GAP — top 500 keywords"),
    ("material_demand.csv",      "SC material demand"),
]

# numeric columns -> formatting
NUM_COLS = {"search_volume", "etv", "position", "keyword_difficulty", "cpc",
            "keywords_in_space", "etv_in_space", "traffic_share_pct",
            "avg_position", "visibility", "competitors_ranking",
            "sc_current_position", "keyword_count", "total_search_volume",
            "expected_traffic_pos3", "avg_keyword_difficulty",
            "Monthly Search Volume", "Expected Traffic @ #3",
            "priority_score", "Priority Score", "head_keyword_volume",
            "Traffic Potential @#3 (ceiling)", "Realistic Traffic /mo",
            "Hero Volume", "SC Current Position"}
INT_COLS = {"search_volume", "position", "keyword_difficulty",
            "keywords_in_space", "etv_in_space", "competitors_ranking",
            "sc_current_position", "keyword_count", "total_search_volume",
            "expected_traffic_pos3", "avg_keyword_difficulty",
            "Monthly Search Volume", "Expected Traffic @ #3",
            "priority_score", "Priority Score", "head_keyword_volume",
            "Traffic Potential @#3 (ceiling)", "Realistic Traffic /mo",
            "Hero Volume", "SC Current Position"}


def add_sheet(wb, csv_name, tab_name, first):
    ws = wb.active if first else wb.create_sheet()
    ws.title = tab_name

    with (DATA_DIR / csv_name).open() as f:
        reader = csv.reader(f)
        header = next(reader)
        ws.append(header)
        col_idx = {name: i for i, name in enumerate(header)}
        for row in reader:
            out = []
            for i, val in enumerate(row):
                name = header[i]
                if name in NUM_COLS and val not in ("", None):
                    try:
                        num = float(val)
                        out.append(int(round(num)) if name in INT_COLS else round(num, 2))
                        continue
                    except ValueError:
                        pass
                out.append(val)
            ws.append(out)

    # header styling
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # column widths
    widths = {
        "keyword": 38, "search_volume": 14, "etv": 12, "position": 10,
        "ranking_url": 60, "keyword_difficulty": 18, "cpc": 8, "search_intent": 15,
        "competition": 13, "source": 16,
        "domain": 32, "category": 14, "keywords_in_space": 17,
        "etv_in_space": 14, "traffic_share_pct": 16, "avg_position": 12,
        "visibility": 12,
        "competitors_ranking": 18, "competitor_list": 34,
        "sc_current_position": 18, "theme": 22, "recommended_action": 22,
        "target_page_type": 16, "target_handle_or_url": 34,
        "target_page_title": 26, "target_url": 34, "keyword_count": 14,
        "total_search_volume": 18, "example_keywords": 60,
        "expected_traffic_pos3": 20, "avg_keyword_difficulty": 20,
        "winnability": 12, "sc_product_line": 32,
        "Page": 26, "New / Optimise": 14, "SC Product Line": 32,
        "Current / Proposed URL": 48,
        "To Rank For": 70, "Monthly Search Volume": 20,
        "Expected Traffic @ #3": 20, "Winnability": 12,
        "top_keywords": 70, "priority_score": 14, "Priority Score": 14,
        "head_keyword_volume": 18, "Traffic Potential @#3 (ceiling)": 26,
        "Winnability (SERP)": 16, "SERP Reality": 52, "Realistic Traffic /mo": 20,
        "Page to Beat": 60, "Hero Keyword": 30, "Hero Volume": 12,
        "Supporting Keywords": 60, "Plan Status": 26, "Page to Beat Rank": 36,
        "Target Hero Keyword": 30, "Current / New Page": 48, "Page Type": 11,
        "SC Current Position": 18,
    }
    for name, idx in col_idx.items():
        ws.column_dimensions[get_column_letter(idx + 1)].width = widths.get(name, 14)

    print(f"  {tab_name}: {ws.max_row - 1} rows")


def main():
    wb = Workbook()
    for i, (csv_name, tab_name) in enumerate(TABS):
        add_sheet(wb, csv_name, tab_name, first=(i == 0))
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
